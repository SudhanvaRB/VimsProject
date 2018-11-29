'''
Created on Nov 14, 2018

@author: subharad
'''
#from Pytest.test_locators import Locators
from test_locators import Locators
import time
import openpyxl
#from selenium.common.exceptions import NoSuchElementException

class Login():
    
    def __init__(self,driver):
        self.driver = driver
        

    def login(self):
        self.driver.get("http://10.100.245.80:92/VMS/login")
        #wk = openpyxl.load_workbook()
        self.driver.find_element_by_id(Locators.username_text_box).send_keys("tihollan")
        self.driver.find_element_by_id(Locators.login_button).click()
        time.sleep(1)
    
    def param_login(self):
        wk = openpyxl.load_workbook("C:/selenium_scripts/Selenium/driver/login_data.xlsx")
        act = wk.active
        rws = act.max_row
        #print(rws)
        col = act.max_column
        #print(col)
        for i in range(1,rws+1):
            for j in range(1,col+1):
                c=act.cell(i,j)
                self.driver.get("http://10.100.245.80:92/VMS/login")
                #wk = openpyxl.load_workbook()
                self.driver.find_element_by_id(Locators.username_text_box).send_keys(c.value)
                time.sleep(1)
                self.driver.find_element_by_id(Locators.login_button).click()
                time.sleep(1)